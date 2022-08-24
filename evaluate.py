# -*- coding: utf-8 -*-
"""
Created on Sat May  4 15:00:43 2019

@author: lijun
"""

import argparse
from datetime import datetime
from random import shuffle
import random
import os
import sys
import time
import math
import numpy as np
import glob
from PIL import Image
from gdaldiy import *
import openpyxl
from collections import OrderedDict
parser = argparse.ArgumentParser(description='')
parser.add_argument("--out_dir", default='./test_output/',help="Output Folder") #保存x域的输入图片与生成的y域图片的路径
parser.add_argument("--snowmask_dir", default='f:/lijun/data/graduatedata/clouddetection/S2A/test/snowmask/',help="Output Folder")
parser.add_argument("--label_data_path", default=r'F:\lijun\data\graduatedata\clouddetection\La8\e2etest\label\\', help="path of x training datas.") #x域的训练图片路径 
args = parser.parse_args()
rgb_colors=OrderedDict([
    ("nodata",np.array([0],dtype=np.uint8)),
    ("cloud-free",np.array([128],dtype=np.uint8)),
    ("cloud",np.array([255],dtype=np.uint8))]) 
label_threshold=128.
mask_threshold=0.5
def make_test_data_list(x_data_path): #make_test_data_list函数得到测试中的x域和y域的图像路径名称列表	
    x_input_images = glob.glob(os.path.join(x_data_path, "*")) #读取全部的x域图像路径名称列表
    return x_input_images 
def read_testimg(datapath):
    img=imgread(datapath)
    img=img[np.newaxis,:]
    return img

def write_results2xlsx(classnames,filenames,metrics,savename,mean=False):
    batch=metrics.shape[0]
    Apnames=['OA','Precision','Recall',"F1-score","IOU"] 
    savefile=args.out_dir+os.getcwd().split('\\')[-1]+savename+'.xlsx'
    if os.path.exists(savefile): #如果保存x域测试结果的文件夹不存在则创建
        wb=openpyxl.load_workbook(savefile)
    else:
        wb=openpyxl.Workbook()
        sheet1=wb.active
        sheet1.title=classnames[0]     
   
    for i in range(len(classnames)):
        sheetsnames=wb.sheetnames
        if classnames[i] not in sheetsnames:
            ws=wb.create_sheet(classnames[i])
        else:
            ws=wb[classnames[i]]
        for j in range(len(Apnames)):
                ws.cell(row=1,column=2+j).value=Apnames[j]
        if mean:
            ws.cell(row=2,column=1).value='mean'                    
            for j in range(len(Apnames)):
                ws.cell(row=2,column=2+j).value=np.mean(metrics,0)[i,j]
        old_nrows=ws.max_row
        for j in range(batch):
            ws.cell(row=old_nrows+1+j,column=1).value=str(filenames[j])
            for k in range(len(Apnames)):        
                ws.cell(row=old_nrows+1+j,column=2+k).value=metrics[j,i,k]
           
    wb.save(savefile)
def evaluate(modelname):
    print('开始精度评价')
    if not os.path.exists(args.out_dir): #如果保存训练中可视化输出的文件夹不存在则创建
        os.makedirs(args.out_dir)
    roc_wb=openpyxl.load_workbook('./imglist.xlsx')
    test_sheet=roc_wb['WHUS2-CDtest']
    label_data_path=test_sheet.cell(row=1,column=3).value
    print(label_data_path)
    mask_datalists=[]
    for i in range(test_sheet.max_row):
         mask_datalists.append(glob.glob(os.path.join(args.out_dir, modelname+"*"))[0]+'\\'+test_sheet.cell(row=i+1,column=1).value+'.tif')
    filenames=[]
    classnames=['cloud']
    results=np.zeros(shape=(len(mask_datalists),len(classnames),5))
    meanresults=np.zeros(shape=(1,len(classnames),5))
    labels=[]
    masks=[]
    image_size=10980
    for i in range(len(mask_datalists)):
        label_zero=np.zeros(shape=(image_size,image_size),dtype=np.uint8)
        Mask_zero=np.zeros(shape=(image_size,image_size),dtype=np.uint8)
        maskname=mask_datalists[i].split('\\')[-1].split('.')[0]
        filenames.append(maskname)
        labelpath=label_data_path+maskname+'.tif'
        label=imgread(labelpath)
        label_zero[:label.shape[0],:label.shape[1]]=label
        labels.append(label_zero)
        Mask=imgread(mask_datalists[i])    
        Mask_zero[:Mask.shape[0],:Mask.shape[1]]=Mask
        masks.append(Mask_zero) 
        A,P,R,F,I = single_evaluate(label,Mask,192)
        lists=[A,P,R,F,I]
        for k in range(len(lists)):
            results[i,0,k]=lists[k]
    write_results2xlsx(classnames,filenames,results,modelname)
    A,P,R,F,I = single_evaluate(np.concatenate(labels,0),np.concatenate(masks,0),192)
    mean_lists=[A,P,R,F,I]
    for m in range(len(classnames)):
        for n in range(len(lists)):
            meanresults[0,m,n]=mean_lists[n]
    write_results2xlsx(['allmean'],['allmean'],meanresults,modelname)             
    print('评价完毕，检测精度结果已保存')
    print('开始保存检测结果')

def evaluateBiome(modelname,test_sheet):
    if not os.path.exists(args.out_dir): #如果保存训练中可视化输出的文件夹不存在则创建
        os.makedirs(args.out_dir)    
    mask_datalists=[]
    for i in range(test_sheet.max_row-1):
        mask_datalists.append(glob.glob(os.path.join(args.out_dir, modelname+"*"))[0]+'\\'+test_sheet.cell(row=i+1,column=1).value+'.jpg')
    
    labels=[]
    masks=[]
    image_size=10000
    for i in range(len(mask_datalists)):
        label_zero=np.zeros(shape=(image_size,image_size),dtype=np.uint8)
        Mask_zero=np.zeros(shape=(image_size,image_size),dtype=np.uint8)
        maskname=mask_datalists[i].split('\\')[-1].split('.')[0]
        labelpath=args.label_data_path+maskname+'.tif'
        label=imgread(labelpath)
        label_zero[:label.shape[0],:label.shape[1]]=label
        labels.append(label_zero)
        Mask=imgread(mask_datalists[i])
        Mask_zero[:Mask.shape[0],:Mask.shape[1]]=Mask
        masks.append(Mask_zero)       
    A,P,R,F,I = single_evaluate(np.concatenate(labels,0),np.concatenate(masks,0),192)
    return [A,P,R,F,I]        
def multi_evaluate(modelname):
    print('开始精度评价')
    if not os.path.exists(args.out_dir): #如果保存训练中可视化输出的文件夹不存在则创建
        os.makedirs(args.out_dir)
    roc_wb=openpyxl.load_workbook('./landcovers.xlsx')
    namelists=roc_wb.sheetnames
    results=np.zeros(shape=(len(namelists),1,5))
    for i in range(len(namelists)):
        sheetname=namelists[i]
        test_sheet=roc_wb[sheetname]
        l=evaluateBiome(modelname,test_sheet)
        for k in range(len(l)):
            results[i,0,k]=l[k]
        print(sheetname+',评价完毕')
    write_results2xlsx(['cloud'],namelists,results,'landcoverdone')         
    print('评价完毕，检测精度结果已保存')
    print('开始保存检测结果') 
def single_evaluate(input_label,input_Mask,k):
    # label=(label>=label_threshold).astype(int)
    # Mask=(Mask>=mask_threshold).astype(int)
    ones=(input_label>=0).astype(np.uint8)
    label=np.asarray((input_label>=k).astype(np.uint8))*ones
    Mask=np.asarray((input_Mask>=k).astype(np.uint8))*ones
    
    TP=np.sum(Mask*label)
    TN=np.sum((ones-Mask)*(ones-label))
    Accuracy=(TP+TN+1e-20)/(np.sum(ones)+1e-20)
    Precision=(TP+1e-20)/(np.sum(Mask)+1e-20)
    Recall=(TP+1e-20)/(np.sum(label)+1e-20)
    F1=2*(Precision*Recall)/(Precision+Recall+1e-20)
    iou=(np.sum(Mask*label)+1e-20)/(np.sum(((Mask+label)>0).astype(np.uint8))+1e-20) 
    return Accuracy,Precision,Recall,F1,iou   



if __name__ == '__main__':
    evaluate("mask")
    # evaluateall("mask12")
    # multi_evaluate("mask12")
    #evaluate_tf(str(30000))

